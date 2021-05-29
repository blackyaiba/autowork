import os
import openpyxl as op

creat_path = 'C:\\Users\\shilanqian\\Desktop\\5月结关\\'
input_file = 'C:\\Users\\shilanqian\\Desktop\\5月结关\\python通关.xlsx'
jieguanmoban_file = r'\\192.168.128.27\DFQD-Shared\市场服务部\物流组\关务\（日常）强行结关（目前每月下旬提交之前数据）\模板\结关申请模板.docx'


inputfileworkbook = op.load_workbook(filename=input_file)
inputsheet = inputfileworkbook['通关明细表']
needcreat = inputsheet['I']

# 创建文件夹
for cell in needcreat:
    namexuhao = inputsheet.cell(row=cell.row, column=1)
    namebgd = inputsheet.cell(row=cell.row, column=5)
    namefapiao = inputsheet.cell(row=cell.row, column=6)
    namexiangzhu = inputsheet.cell(row=cell.row, column=13)
    newdocumentname = str(namexuhao.value) + " " + str(namebgd.value) + " " + str(namefapiao.value) + " " + str(
        namexiangzhu.value)
    # print(cell.row,cell.column)
    wenjianjiaifexist = os.path.exists(creat_path + newdocumentname)

    if cell.value == '5月制作':
        print('需创建文件夹：' + newdocumentname)


        if not wenjianjiaifexist:
            os.makedirs(creat_path + newdocumentname)
            print('文件夹创建完毕')

        else:
            print("文件夹已存在")
           111
